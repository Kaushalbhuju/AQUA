import datetime
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q, Max, Count
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dashboard.models import Student
from .models import ExamResult, SkillExamResult


@login_required
def exam_selection(request):
    return render(request, 'japanese_exam/exam_selection.html')


@login_required
def exam_student_list(request, exam_type):
    if exam_type not in ['JFT', 'JLPT']:
        messages.error(request, 'Invalid exam type.')
        return redirect('japanese_exam:exam_selection')

    search_query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    attempt_filter = request.GET.get('attempt_filter', 'all')

    students_qs = Student.objects.filter(status='approved').order_by('full_name')
    if search_query:
        students_qs = students_qs.filter(
            Q(full_name__icontains=search_query) |
            Q(student_id__icontains=search_query) |
            Q(phone__icontains=search_query)
        )

    all_results = ExamResult.objects.filter(
        exam_type=exam_type
    ).select_related('student', 'recorded_by').order_by('student_id', 'attempt_number')

    # Group attempts by student
    student_attempts_map = {}
    for r in all_results:
        if r.student_id not in student_attempts_map:
            student_attempts_map[r.student_id] = []
        student_attempts_map[r.student_id].append(r)

    student_data = []
    passed_students_count = 0
    attempted_students_count = 0

    all_approved_students = Student.objects.filter(status='approved')
    total_approved_count = all_approved_students.count()

    for s in students_qs:
        attempts = student_attempts_map.get(s.id, [])
        attempt_count = len(attempts)
        latest_attempt = attempts[-1] if attempts else None
        
        # Check if passed in any attempt
        pass_attempt = next((a for a in reversed(attempts) if a.status == 'pass'), None)
        has_passed = pass_attempt is not None

        if attempts:
            attempted_students_count += 1
            if has_passed:
                passed_students_count += 1

        overall_status = 'pass' if has_passed else ('fail' if attempts else 'none')

        # Filter by status
        if status_filter == 'pass' and not has_passed:
            continue
        if status_filter == 'fail' and (has_passed or not attempts):
            continue
        if status_filter == 'pending' and attempts:
            continue

        # Filter by attempts
        if attempt_filter == 'single' and attempt_count != 1:
            continue
        if attempt_filter == 'multiple' and attempt_count < 2:
            continue
        if attempt_filter == 'none' and attempt_count != 0:
            continue

        student_data.append({
            'student': s,
            'attempts': attempts,
            'attempt_count': attempt_count,
            'next_attempt_number': attempt_count + 1,
            'latest_attempt': latest_attempt,
            'pass_attempt': pass_attempt,
            'has_passed': has_passed,
            'overall_status': overall_status,
        })

    # Overall system metrics for exam_type
    total_attempts_recorded = all_results.count()
    failed_students_count = attempted_students_count - passed_students_count
    pass_rate = round((passed_students_count / attempted_students_count * 100), 1) if attempted_students_count > 0 else 0.0

    today_str = timezone.now().strftime('%Y-%m-%d')

    return render(request, 'japanese_exam/exam_student_list.html', {
        'student_data': student_data,
        'exam_type': exam_type,
        'search_query': search_query,
        'status_filter': status_filter,
        'attempt_filter': attempt_filter,
        'total_approved_count': total_approved_count,
        'attempted_students_count': attempted_students_count,
        'total_attempts_recorded': total_attempts_recorded,
        'passed_students_count': passed_students_count,
        'failed_students_count': failed_students_count,
        'pass_rate': pass_rate,
        'today_str': today_str,
    })


@login_required
def record_exam_attempt(request, exam_type):
    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        status_val = request.POST.get('status')
        score_val = request.POST.get('score')
        exam_date_val = request.POST.get('exam_date')
        remarks_val = request.POST.get('remarks', '').strip()
        attempt_num_val = request.POST.get('attempt_number')

        if student_id and status_val:
            student = get_object_or_404(Student, id=student_id, status='approved')

            existing_count = ExamResult.objects.filter(student=student, exam_type=exam_type).count()
            attempt_number = int(attempt_num_val) if attempt_num_val and attempt_num_val.isdigit() else existing_count + 1

            parsed_date = timezone.now().date()
            if exam_date_val:
                try:
                    parsed_date = datetime.datetime.strptime(exam_date_val, '%Y-%m-%d').date()
                except ValueError:
                    pass

            score_int = int(score_val) if score_val and score_val.isdigit() else None

            ExamResult.objects.create(
                student=student,
                exam_type=exam_type,
                attempt_number=attempt_number,
                exam_date=parsed_date,
                status=status_val,
                score=score_int,
                remarks=remarks_val,
                recorded_by=request.user
            )
            messages.success(request, f"Recorded Attempt #{attempt_number} ({status_val.upper()}) for {student.full_name}.")

    return redirect('japanese_exam:exam_student_list', exam_type=exam_type)


@login_required
def edit_exam_attempt(request, result_id):
    result = get_object_or_404(ExamResult, id=result_id)
    exam_type = result.exam_type

    if request.method == 'POST':
        status_val = request.POST.get('status')
        score_val = request.POST.get('score')
        exam_date_val = request.POST.get('exam_date')
        remarks_val = request.POST.get('remarks', '').strip()
        attempt_num_val = request.POST.get('attempt_number')

        if status_val:
            result.status = status_val
            result.score = int(score_val) if score_val and score_val.isdigit() else None
            result.remarks = remarks_val
            if attempt_num_val and attempt_num_val.isdigit():
                result.attempt_number = int(attempt_num_val)

            if exam_date_val:
                try:
                    result.exam_date = datetime.datetime.strptime(exam_date_val, '%Y-%m-%d').date()
                except ValueError:
                    pass

            result.recorded_by = request.user
            result.save()
            messages.success(request, f"Updated Attempt #{result.attempt_number} for {result.student.full_name}.")

    return redirect('japanese_exam:exam_student_list', exam_type=exam_type)


@login_required
def delete_exam_attempt(request, result_id):
    result = get_object_or_404(ExamResult, id=result_id)
    exam_type = result.exam_type
    student_name = result.student.full_name
    attempt_no = result.attempt_number

    if request.method == 'POST':
        result.delete()
        messages.success(request, f"Deleted Attempt #{attempt_no} record for {student_name}.")

    return redirect('japanese_exam:exam_student_list', exam_type=exam_type)


@login_required
def export_exam_excel(request, exam_type):
    if exam_type not in ['JFT', 'JLPT']:
        messages.error(request, 'Invalid exam type.')
        return redirect('japanese_exam:exam_selection')

    wb = openpyxl.Workbook()
    
    # Setup styles
    font_title = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    font_subtitle = Font(name='Calibri', size=11, bold=True, color='1E3A8A')
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_bold = Font(name='Calibri', size=11, bold=True)
    font_regular = Font(name='Calibri', size=11)
    font_pass = Font(name='Calibri', size=11, bold=True, color='065F46')
    font_fail = Font(name='Calibri', size=11, bold=True, color='991B1B')

    fill_navy = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
    fill_sub_blue = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    fill_pass = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
    fill_fail = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
    fill_zebra = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
    fill_gray = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

    thin_border_side = Side(style='thin', color='D1D5DB')
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Fetch data
    students = Student.objects.filter(status='approved').order_by('full_name')
    results = ExamResult.objects.filter(exam_type=exam_type).select_related('student', 'recorded_by').order_by('student__full_name', 'attempt_number')

    # -------------------------------------------------------------
    # SHEET 1: ALL EXAM ATTEMPTS HISTORY
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = f"{exam_type} Attempts Log"
    ws1.views.sheetView[0].showGridLines = True

    # Header title
    ws1.merge_cells('A1:K1')
    cell_a1 = ws1['A1']
    cell_a1.value = f"AQUA EDUCATION - {exam_type} EXAM ALL ATTEMPTS REPORT"
    cell_a1.font = font_title
    cell_a1.fill = fill_navy
    cell_a1.alignment = align_center
    ws1.row_dimensions[1].height = 40

    # Subtitle
    ws1.merge_cells('A2:K2')
    cell_a2 = ws1['A2']
    cell_a2.value = f"Generated On: {timezone.now().strftime('%Y-%m-%d %H:%M')} | Total Attempts Recorded: {results.count()}"
    cell_a2.font = font_subtitle
    cell_a2.fill = fill_sub_blue
    cell_a2.alignment = align_center
    ws1.row_dimensions[2].height = 25

    # Headers at row 4
    headers1 = [
        "S.N.", "Student ID", "Full Name", "Phone",
        "Attempt #", "Exam Date", "Score", "Status",
        "Remarks", "Recorded By", "Recorded Date"
    ]
    
    ws1.row_dimensions[4].height = 28
    for col_num, header in enumerate(headers1, 1):
        cell = ws1.cell(row=4, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_all

    row_idx = 5
    for idx, r in enumerate(results, 1):
        ws1.row_dimensions[row_idx].height = 22
        
        status_text = r.get_status_display().upper() if r.status else "PENDING"
        exam_date_str = r.exam_date.strftime('%Y-%m-%d') if r.exam_date else "-"
        recorded_by_str = r.recorded_by.get_full_name() if r.recorded_by and r.recorded_by.get_full_name() else (r.recorded_by.username if r.recorded_by else "-")
        created_str = r.created_at.strftime('%Y-%m-%d') if r.created_at else "-"

        data_row = [
            idx,
            r.student.student_id or "-",
            r.student.full_name,
            r.student.phone or "-",
            f"Attempt {r.attempt_number}",
            exam_date_str,
            r.score if r.score is not None else "-",
            status_text,
            r.remarks or "-",
            recorded_by_str,
            created_str
        ]

        row_fill = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)

        for col_idx, val in enumerate(data_row, 1):
            c = ws1.cell(row=row_idx, column=col_idx, value=val)
            c.font = font_regular
            c.border = border_all
            c.fill = row_fill
            
            if col_idx in [1, 2, 5, 6, 7, 10, 11]:
                c.alignment = align_center
            else:
                c.alignment = align_left

            # Highlight status column (column 8)
            if col_idx == 8:
                if r.status == 'pass':
                    c.fill = fill_pass
                    c.font = font_pass
                elif r.status == 'fail':
                    c.fill = fill_fail
                    c.font = font_fail

        row_idx += 1

    # Auto-fit columns Sheet 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # -------------------------------------------------------------
    # SHEET 2: STUDENT CANDIDATE SUMMARY
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title=f"{exam_type} Candidate Summary")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells('A1:I1')
    cell_b1 = ws2['A1']
    cell_b1.value = f"AQUA EDUCATION - {exam_type} CANDIDATE SUMMARY REPORT"
    cell_b1.font = font_title
    cell_b1.fill = fill_navy
    cell_b1.alignment = align_center
    ws2.row_dimensions[1].height = 40

    ws2.merge_cells('A2:I2')
    cell_b2 = ws2['A2']
    cell_b2.value = f"Total Approved Candidates: {students.count()} | Report Date: {timezone.now().strftime('%Y-%m-%d')}"
    cell_b2.font = font_subtitle
    cell_b2.fill = fill_sub_blue
    cell_b2.alignment = align_center
    ws2.row_dimensions[2].height = 25

    headers2 = [
        "S.N.", "Student ID", "Full Name", "Phone",
        "Total Attempts", "Latest Attempt Date", "Best Score", "Passed?", "Final Status"
    ]
    
    ws2.row_dimensions[4].height = 28
    for col_num, header in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=col_num, value=header)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_all

    row_idx = 5
    for idx, s in enumerate(students, 1):
        s_attempts = [r for r in results if r.student_id == s.id]
        total_att = len(s_attempts)
        
        pass_att = next((a for a in reversed(s_attempts) if a.status == 'pass'), None)
        latest_att = s_attempts[-1] if s_attempts else None

        scores = [a.score for a in s_attempts if a.score is not None]
        best_score = max(scores) if scores else "-"
        latest_date = latest_att.exam_date.strftime('%Y-%m-%d') if (latest_att and latest_att.exam_date) else "-"

        if pass_att:
            final_status = "PASSED"
            is_passed_str = "YES"
        elif s_attempts:
            final_status = "FAILED / RETRYING"
            is_passed_str = "NO"
        else:
            final_status = "NOT ATTEMPTED"
            is_passed_str = "-"

        ws2.row_dimensions[row_idx].height = 22
        data_row = [
            idx,
            s.student_id or "-",
            s.full_name,
            s.phone or "-",
            total_att,
            latest_date,
            best_score,
            is_passed_str,
            final_status
        ]

        row_fill = fill_zebra if row_idx % 2 == 0 else PatternFill(fill_type=None)

        for col_idx, val in enumerate(data_row, 1):
            c = ws2.cell(row=row_idx, column=col_idx, value=val)
            c.font = font_regular
            c.border = border_all
            c.fill = row_fill

            if col_idx in [1, 2, 5, 6, 7, 8, 9]:
                c.alignment = align_center
            else:
                c.alignment = align_left

            if col_idx == 9:
                if final_status == "PASSED":
                    c.fill = fill_pass
                    c.font = font_pass
                elif final_status == "FAILED / RETRYING":
                    c.fill = fill_fail
                    c.font = font_fail
                else:
                    c.fill = fill_gray

        row_idx += 1

    # Auto-fit columns Sheet 2
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{exam_type}_Exam_Report_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def skill_exam_student_list(request):
    students = Student.objects.filter(status='approved').order_by('full_name')

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        skill_category = request.POST.get('skill_category')
        status_val = request.POST.get('status')
        score_val = request.POST.get('score')

        if student_id and skill_category and status_val:
            student = get_object_or_404(Student, id=student_id, status='approved')
            SkillExamResult.objects.update_or_create(
                student=student,
                skill_category=skill_category,
                defaults={
                    'status': status_val,
                    'score': int(score_val) if score_val else None,
                    'recorded_by': request.user,
                }
            )
            messages.success(request, f'Skill result recorded for {student.full_name}')
        return redirect('japanese_exam:skill_exam_student_list')

    results_map = {
        r.student_id: r
        for r in SkillExamResult.objects.all().select_related('student')
    }

    student_data = []
    for s in students:
        result = results_map.get(s.id)
        student_data.append({
            'student': s,
            'result': result,
        })

    return render(request, 'japanese_exam/skill_exam_student_list.html', {
        'student_data': student_data,
        'skill_categories': SkillExamResult.SKILL_CATEGORY_CHOICES,
    })
