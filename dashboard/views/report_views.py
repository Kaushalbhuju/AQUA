import calendar
from datetime import date
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils.translation import gettext as _
from dashboard.decorators import check_role
from django.http import HttpResponse


def localized_classroom_name(name):
    mapping = {
        'Class A': _('Class A'),
        'Class B': _('Class B'),
        'Class C': _('Class C'),
        'Class D': _('Class D'),
    }
    return mapping.get(name, name)

@login_required(login_url='login')
@check_role('teacher')
def teacher_report(request):
    """View indicating student attendance and monthly notes together"""
    from dashboard.models import Classroom, ClassStudent, AttendanceRecord, StudentDailyNote
    
    classrooms = Classroom.objects.filter(teacher=request.user)
    
    classroom_id = request.GET.get('classroom_id')
    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        year = today.year
        month = today.month
        
    selected_classroom = None
    students = []
    student_data = []
    
    if not classrooms.exists():
        # User has no classrooms
        pass
    else:
        if not classroom_id:
            selected_classroom = classrooms.first()
        else:
            try:
                selected_classroom = classrooms.get(id=classroom_id)
            except Classroom.DoesNotExist:
                selected_classroom = classrooms.first()
                
        if selected_classroom:
            class_students = ClassStudent.objects.filter(classroom=selected_classroom).select_related('student')
            students = [cs.student for cs in class_students]
            
            days_in_month = calendar.monthrange(year, month)[1]
            days = list(range(1, days_in_month + 1))
            
            for student in students:
                records = AttendanceRecord.objects.filter(
                    student=student,
                    classroom=selected_classroom,
                    attendance_date__year=year,
                    attendance_date__month=month
                )
                attendance_map = {r.attendance_date.day: r.status for r in records}
                
                present_count = 0
                holiday_count = 0
                for day in days:
                    status = attendance_map.get(day, '')
                    if status == 'present':
                        present_count += 1
                    elif status == 'holiday':
                        holiday_count += 1
                        
                effective_days = days_in_month - holiday_count
                percentage = round((present_count / effective_days) * 100) if effective_days > 0 else 0
                
                notes = StudentDailyNote.objects.filter(
                    student=student,
                    note_date__year=year,
                    note_date__month=month
                ).order_by('-note_date')
                
                student_data.append({
                    'student': student,
                    'percentage': percentage,
                    'notes': notes,
                    'present_count': present_count,
                    'effective_days': effective_days
                })

    for classroom in classrooms:
        classroom.localized_name = localized_classroom_name(classroom.name)
            
    month_name = calendar.month_name[month]
    
    # Pre-calculate prev/next month for navigation
    if month == 1:
        prev_month, prev_year = 12, year - 1
    else:
        prev_month, prev_year = month - 1, year
        
    if month == 12:
        next_month, next_year = 1, year + 1
    else:
        next_month, next_year = month + 1, year

    context = {
        'classrooms': classrooms,
        'selected_classroom': selected_classroom,
        'selected_classroom_display': localized_classroom_name(selected_classroom.name) if selected_classroom else None,
        'student_data': student_data,
        'month': month,
        'year': year,
        'month_name': month_name,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        'page_title': _('Class Report')
    }
    
    return render(request, 'dashboards/teacher_report.html', context)


@login_required(login_url='login')
@check_role('teacher')
def export_report_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from dashboard.models import Classroom, AttendanceRecord, StudentDailyNote, ClassStudent
    
    classroom_id = request.GET.get('classroom_id')
    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        year = today.year
        month = today.month

    if not classroom_id:
        messages.error(request, _('Please select a classroom before exporting.'))
        return redirect('dashboard:teacher_report')

    selected_classroom = Classroom.objects.filter(id=classroom_id, teacher=request.user).first()
    if not selected_classroom:
        messages.error(request, _('Selected classroom was not found.'))
        return redirect('dashboard:teacher_report')
    class_students = ClassStudent.objects.filter(classroom=selected_classroom).select_related('student')
    students = [cs.student for cs in class_students]
    
    days_in_month = calendar.monthrange(year, month)[1]
    days = list(range(1, days_in_month + 1))
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_name[month][:3]} {year} Report"
    
    # Headers
    headers = ['S.N', 'Full Name', 'Japanese Name', 'Gender', 'Age', 'Attendance %', 'Notes Summary']
    ws.append(headers)
    
    # Style building
    header_font = Font(bold=True, color='FFFFFF')
    fill = PatternFill(start_color='004E89', end_color='004E89', fill_type='solid')
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['G'].width = 50
    
    for idx, student in enumerate(students, 1):
        records = AttendanceRecord.objects.filter(
            student=student, classroom=selected_classroom,
            attendance_date__year=year, attendance_date__month=month
        )
        attendance_map = {r.attendance_date.day: r.status for r in records}
        present_count = 0
        holiday_count = 0
        for day in days:
            status = attendance_map.get(day, '')
            if status == 'present': present_count += 1
            elif status == 'holiday': holiday_count += 1
                
        effective_days = days_in_month - holiday_count
        pct = f"{round((present_count / effective_days) * 100)}%" if effective_days > 0 else "0%"
        
        notes = StudentDailyNote.objects.filter(
            student=student, note_date__year=year, note_date__month=month
        ).order_by('-note_date')
        
        notes_str = "\n".join([f"{n.note_date.strftime('%d')}: {n.content}" for n in notes])
        
        row_data = [
            idx,
            student.full_name,
            student.japanese_name or '',
            student.gender or '',
            student.calculated_age or student.age,
            pct,
            notes_str
        ]
        
        ws.append(row_data)
        
        # wrap text for notes
        ws.cell(row=idx+1, column=7).alignment = Alignment(wrap_text=True, vertical='top')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Report_{selected_classroom.name}_{month}_{year}.xlsx"'
    wb.save(response)
    
    return response

@login_required(login_url='login')
@check_role('teacher')
def export_report_pdf(request):
    from dashboard.models import Classroom, ClassStudent, AttendanceRecord, StudentDailyNote
    from django.template.loader import render_to_string
    import xhtml2pdf.pisa as pisa
    from io import BytesIO
    
    classroom_id = request.GET.get('classroom_id')
    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        year = today.year
        month = today.month

    if not classroom_id:
        messages.error(request, _('Please select a classroom before exporting.'))
        return redirect('dashboard:teacher_report')

    selected_classroom = Classroom.objects.filter(id=classroom_id, teacher=request.user).first()
    if not selected_classroom:
        messages.error(request, _('Selected classroom was not found.'))
        return redirect('dashboard:teacher_report')
    class_students = ClassStudent.objects.filter(classroom=selected_classroom).select_related('student')
    students = [cs.student for cs in class_students]
    
    days_in_month = calendar.monthrange(year, month)[1]
    days = list(range(1, days_in_month + 1))
    
    student_data = []
    for student in students:
        records = AttendanceRecord.objects.filter(
            student=student, classroom=selected_classroom,
            attendance_date__year=year, attendance_date__month=month
        )
        attendance_map = {r.attendance_date.day: r.status for r in records}
        present_count = 0
        holiday_count = 0
        for day in days:
            status = attendance_map.get(day, '')
            if status == 'present': present_count += 1
            elif status == 'holiday': holiday_count += 1
                
        effective_days = days_in_month - holiday_count
        percentage = round((present_count / effective_days) * 100) if effective_days > 0 else 0
        
        notes = StudentDailyNote.objects.filter(
            student=student, note_date__year=year, note_date__month=month
        ).order_by('-note_date')
        
        student_data.append({
            'student': student,
            'percentage': percentage,
            'notes': notes,
            'present_count': present_count,
            'effective_days': effective_days,
        })
        
    context = {
        'classroom': selected_classroom,
        'classroom_display_name': localized_classroom_name(selected_classroom.name),
        'student_data': student_data,
        'month_name': calendar.month_name[month],
        'year': year
    }
    
    html = render_to_string('dashboards/teacher_report_pdf.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Report_{selected_classroom.name}_{month}_{year}.pdf"'
    
    # Create PDF
    pisa_status = pisa.CreatePDF(
        html, dest=response
    )
    
    if pisa_status.err:
        return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response
